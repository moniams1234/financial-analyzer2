"""
export_utils.py
Builds Excel workbook and JSON export.

Excel sheets:
  Raw_Trial_Balance, Mapping, Mapp, Balance_Sheet, P&L, Red_Flags, Summary
"""
from __future__ import annotations

import io
import json
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ─── helpers ────────────────────────────────────────────────────────────────

def _hdr(ws, row: int, n_cols: int, bg: str = "1E3A5F", fg: str = "FFFFFF"):
    fill = PatternFill("solid", start_color=bg, end_color=bg)
    font = Font(bold=True, color=fg, name="Arial", size=10)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _auto_w(ws):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 55)


def _write_df(ws, df: pd.DataFrame, start: int = 1):
    for ci, col in enumerate(df.columns, 1):
        ws.cell(row=start, column=ci, value=str(col))
    _hdr(ws, start, len(df.columns))
    for ri, row in enumerate(df.itertuples(index=False), start + 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci)
            if isinstance(val, float) and not np.isnan(val):
                cell.value = round(val, 2)
                cell.number_format = "#,##0.00"
            elif isinstance(val, (int, np.integer)):
                cell.value = int(val)
            else:
                cell.value = "" if (val is None or str(val) == "nan") else val


def _section(ws, row: int, text: str):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=12, name="Arial", color="1E3A5F")
    return row + 1


# ─── main export ────────────────────────────────────────────────────────────

def build_excel_export(
    raw_df: pd.DataFrame,
    mapping: dict,
    mapp_df: pd.DataFrame,
    bs: dict,
    pnl: dict,
    flags: list[dict],
    kpis: dict,
    filename: str,
    period: str,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    # ── Raw_Trial_Balance ────────────────────────────────────────────────────
    ws = wb.create_sheet("Raw_Trial_Balance")
    _write_df(ws, raw_df)
    _auto_w(ws)

    # ── Mapping ──────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Mapping")
    map_rows = [{"account_number": k, "side": v["side"], "group": v["group"]}
                for k, v in mapping.items()]
    if map_rows:
        _write_df(ws, pd.DataFrame(map_rows))
        _auto_w(ws)
    else:
        ws.cell(1, 1, "No mapping loaded.")

    # ── Mapp ─────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Mapp")
    if mapp_df is not None and not mapp_df.empty:
        cols = ["account_number", "account_name", "group", "side",
                "debit", "credit", "saldo_dt", "saldo_ct", "persaldo",
                "persaldo_group", "mapping_status"]
        cols = [c for c in cols if c in mapp_df.columns]
        _write_df(ws, mapp_df[cols])
        _auto_w(ws)

    # ── Balance_Sheet ────────────────────────────────────────────────────────
    ws = wb.create_sheet("Balance_Sheet")
    r = 1
    r = _section(ws, r, "ASSETS")
    if bs and not bs["assets_by_group"].empty:
        _write_df(ws, bs["assets_by_group"], start=r)
        r += len(bs["assets_by_group"]) + 1
    ws.cell(r, 1, "Total Assets").font = Font(bold=True, name="Arial")
    ws.cell(r, 2, round(bs.get("total_assets", 0), 2)).number_format = "#,##0.00"
    r += 2
    r = _section(ws, r, "LIABILITIES")
    if bs and not bs["liabilities_by_group"].empty:
        _write_df(ws, bs["liabilities_by_group"], start=r)
        r += len(bs["liabilities_by_group"]) + 1
    ws.cell(r, 1, "Total Liabilities").font = Font(bold=True, name="Arial")
    ws.cell(r, 2, round(bs.get("total_liabilities", 0), 2)).number_format = "#,##0.00"
    r += 1
    diff = bs.get("difference", 0)
    fc = Font(bold=True, name="Arial", color="FF0000" if abs(diff) > 1 else "059669")
    ws.cell(r, 1, "Difference (Assets − Liabilities)").font = fc
    ws.cell(r, 2, round(diff, 2)).font = fc
    _auto_w(ws)

    # ── P&L ──────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("P&L")
    pnl_df = pnl.get("pnl_df", pd.DataFrame())
    if not pnl_df.empty:
        cols = ["account_number", "account_name", "saldo_dt", "saldo_ct", "persaldo_pnl", "pnl_type"]
        cols = [c for c in cols if c in pnl_df.columns]
        _write_df(ws, pnl_df[cols])
        nr = len(pnl_df) + 2
    else:
        ws.cell(1, 1, "No P&L accounts found.")
        nr = 2
    ws.cell(nr, 1, "Net Result").font = Font(bold=True, name="Arial")
    ws.cell(nr, 2, round(pnl.get("net_result", 0), 2)).number_format = "#,##0.00"
    _auto_w(ws)

    # ── Red_Flags ────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Red_Flags")
    if flags:
        _write_df(ws, pd.DataFrame(flags))
        _auto_w(ws)
    else:
        ws.cell(1, 1, "No red flags.")

    # ── Summary ──────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    rows = [
        ("Generated At",      datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Source File",       filename),
        ("Total Accounts",    kpis.get("total_accounts", 0)),
        ("Mapped Accounts",   kpis.get("mapped_accounts", 0)),
        ("Heuristic Mapped",  0),
        ("Total Assets",      round(bs.get("total_assets", 0), 2)),
        ("Total Liabilities", round(bs.get("total_liabilities", 0), 2)),
        ("Total Equity",      round(bs.get("total_liabilities", 0), 2)),   # equity = P side
        ("Balance Difference",round(bs.get("difference", 0), 2)),
        ("Red Flag Errors",   sum(1 for f in flags if f.get("type") == "error")),
        ("Red Flag Warnings", sum(1 for f in flags if f.get("type") == "warning")),
        ("Okres Raportowy",   period),
        ("Wynik Netto",       round(pnl.get("net_result", 0), 2)),
    ]
    bold = Font(bold=True, name="Arial")
    for i, (k, v) in enumerate(rows, 1):
        ws.cell(i, 1, k).font = bold
        c = ws.cell(i, 2, v)
        if isinstance(v, float):
            c.number_format = "#,##0.00"
    _auto_w(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_json_export(
    mapp_df: pd.DataFrame,
    bs: dict,
    pnl: dict,
    flags: list[dict],
    kpis: dict,
    filename: str,
    period: str,
) -> str:
    def _safe_df(df):
        if df is None or (hasattr(df, "empty") and df.empty):
            return []
        return json.loads(df.to_json(orient="records", force_ascii=False, default_handler=str))

    bs_serial: dict = {}
    if bs:
        for k, v in bs.items():
            bs_serial[k] = _safe_df(v) if isinstance(v, pd.DataFrame) else v

    payload = {
        "generated_at": datetime.now().isoformat(),
        "period":        period,
        "source_file":   filename,
        "kpis":          {k: (int(v) if isinstance(v, (np.integer,)) else
                              float(v) if isinstance(v, (np.floating, float)) else v)
                          for k, v in kpis.items()},
        "balance_sheet": bs_serial,
        "pnl_net_result": pnl.get("net_result", 0),
        "red_flags":     flags,
        "mapp":          _safe_df(mapp_df),
    }
    return json.dumps(payload, ensure_ascii=False, indent=2, default=str)
